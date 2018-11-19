import os
import openpyxl
from tkinter.filedialog import askopenfilename


def copy_data(MasterDict, ClientDict, StartRow):  # Create function to copy data from one sheet to another.
    for each in MasterDict:
        MasterCol = MasterDict[each]
        ClientCol = ClientDict[each]
        row = 2
        for AllColumn in range(2, StartRow):
            TempRowClient = ClientWs.cell(row=row, column=ClientCol).value
            TempRowMaster = InfoWs.cell(row=row, column=MasterCol)
            TempRowMaster.value = TempRowClient
            row += 1


def batch_no(ClientName, BatchDate, StartRow):  # Function to create unique core ref.
    BatchID = ClientName + "-" + "2018" + "-" + BatchDate + "-"
    CoreNo = 1
    row = 2
    for AllRow in range(2, StartRow):
        BatchCol = InfoWs.cell(row=row, column=7)
        BatchCol.value = BatchID + str(CoreNo)
        CoreNo += 1
        row += 1


def reinstatement_size(StartRow):  # Function to combine reinstatement length and width
    row = 2
    for AllRow in range(2, StartRow):
        RLength = ClientWs.cell(row=row, column=12).value
        RWidth = ClientWs.cell(row=row, column=13).value
        Rsize = str(RLength) + "x" + str(RWidth)
        TempRow = InfoWs.cell(row=row, column=12)
        TempRow.value = Rsize
        row += 1


def core_tech(Name, StartRow):  # Input coring tech in sheet
    row = 2
    for AllRow in range(2, StartRow):
        TempRow = InfoWs.cell(row=row, column=8)
        TempRow.value = Name
        row += 1


# Dictionaries to match title to column number


Master = {"Address 1": 1,
          "Address 2": 2,
          "Address 3": 3,
          "Address 4": 4,
          "Client Ref.": 6,
          "Company": 9,
          "Road Type": 10,
          "Road Class": 11,
          "Reinstatement Date": 13,
          "StreetWorks": 14}  # Dictionary matching column number with header in master sheet

Knowsley = {"Client Ref.": 1,
            "Address 1": 2,
            "Address 4": 3,
            "Address 2": 4,
            "Address 3": 5,
            "Company": 6,
            "StreetWorks": 7,
            "Road Class": 8,
            "Road Type": 9,
            "Reinstatement Date": 14
            }  # Dictionary matching column number with header in Knowsley sheet

ClientInitial = {"Knowsley": "KN",
                 "North Midlands Construction": "NMC"}  # Dictionary matching client with initial

ClientSheet = askopenfilename()  # ask for location of directory
InfoBank = askopenfilename()

ClientSheet = openpyxl.load_workbook(ClientSheet)  # Load work book
InfoBank = openpyxl.load_workbook(InfoBank)

ClientWs = ClientSheet["Core Log August 2018"]
InfoWs = InfoBank["Client Info"]  # Open correct workbooks and sheets ready to copy

RowCount = ClientWs.max_row
RowCount += 1  # Sets max row in sheet for data copying

ClientName = input("Enter client name: ")
BatchDate = input("Enter batch date (mm): ")
Technician = input("Enter technician full name: ")

ClientName = ClientInitial[ClientName]

batch_no(ClientName, BatchDate, RowCount)  # Runs function with info
reinstatement_size(RowCount)  # Reinstatement Size adding
core_tech(Technician, RowCount)  # Tech
copy_data(Master, Knowsley, RowCount)  # Runs function using dictionary

os.chdir('Users/Ryan McConville/Desktop')

InfoBank.save("Info Bank 2.xlsx")
